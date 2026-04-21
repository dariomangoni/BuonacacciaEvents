from openpyxl import Workbook, load_workbook
import requests
from bs4 import BeautifulSoup
import pandas as pd
import re
import json
from datetime import datetime
from openpyxl.styles import Font, Alignment
from openpyxl.formatting.rule import CellIsRule
from openpyxl.styles import PatternFill
from pathlib import Path
import shutil
from openpyxl.utils.datetime import to_excel
from openpyxl.worksheet.filters import (
    FilterColumn,
    CustomFilters,
    CustomFilter,
)
import os

import argparse
import sys


def clean_data(data_raw):
    try:
        return datetime.strptime(data_raw.strip(), "%d/%m/%Y")
    except:
        return None 
    
def main():
    parser = argparse.ArgumentParser(description="Scraper Eventi BuonaCaccia")
    
    # nargs='?' rende l'argomento posizionale opzionale
    # default="PiccoleOrme" viene usato se non scrivi nulla
    parser.add_argument(
        "tipo", 
        nargs='?', 
        default="PiccoleOrme",
        choices=["PiccoleOrme", "Competenza", "Special"], 
        help="Il tipo di report da generare (default: PiccoleOrme)"
    )

    args = parser.parse_args()

    tipo_eventi = args.tipo
    print(f"Avvio estrazione per: {tipo_eventi}")


    cartella = Path(tipo_eventi)
    cartella.mkdir(parents=True, exist_ok=True)
    shutil.copy('index_template.html', cartella / f'index.html')
    
    json_filename = f"{tipo_eventi}/data.json"

    if os.getenv("GITHUB_ACTIONS") and os.getenv("GITHUB_ACTIONS").lower() == "true":
        json_url = f"https://dariomangoni.github.io/BuonacacciaEvents/{tipo_eventi}/data.json"
        print("Running on cloud.")
        cloud_run = True
    else:
        print("Running locally.")
        cloud_run = False

    # only for filtering; comment to NOT filter
    regioni_filtro = ["Lombardia", "Piemonte", "Veneto", "EmiRo", "Toscana", "Liguria"]
    stato_filtro = ["LIBERO", "CODA"]

    worsheet_title = tipo_eventi
    file_excel = f"{tipo_eventi}/{tipo_eventi}.xlsx"
    if tipo_eventi == "PiccoleOrme":
        buonacaccia_events_url = "https://buonacaccia.net/Events.aspx?RID=&CID=1010101&All=1"
    elif tipo_eventi == "Competenza":
        # la regione dei campetti di competenza viene estratta dal titolo nel quale è riportata "Emilia Romagna" per intero
        regioni_filtro = ["Lombardia", "Piemonte", "Veneto", "Emilia Romagna", "Toscana", "Liguria"]
        buonacaccia_events_url = "https://buonacaccia.net/Events.aspx?RID=&CID=2010104&All=1"
    elif tipo_eventi == "Special":
        buonacaccia_events_url = "https://buonacaccia.net/Events.aspx?RID=&CID=2010101&All=1"
    enable_details_page_scraping = True




    # Scaricamento della pagina
    base_url = "https://buonacaccia.net"
    headers = {'User-Agent': 'Mozilla/5.0'}
    response = requests.get(buonacaccia_events_url, headers=headers)
    soup = BeautifulSoup(response.text, 'html.parser')

    # Trova la tabella principale
    table = soup.find("table", {"id": "MainContent_EventsGridView"})
    tbody = table.find("tbody", recursive=False)
    target = tbody if tbody else table
    rows = target.find_all("tr", recursive=False)

    # Check run precedenti
    try:
        if cloud_run:
            response_json = requests.get(json_url, headers=headers)
            if response_json.status_code == 200:
                print("Previous run data found. Will compare with new data.")
                record_old = response_json.json()
            else:
                print("No previous data found or error fetching it. Will proceed without comparison.")
                record_old = []
        else:
            with open(json_filename, "r", encoding="utf-8") as f:
                record_old = json.load(f)

    except Exception as e:
        print(f"Error fetching previous data: {e}. Will proceed without comparison.")
        record_old = []

    history_loaded = False
    if isinstance(record_old, dict):
        titoli_old = {e['Titolo'] for e in record_old.get("eventi", [])}
        aggiornamento_precedente = record_old.get("aggiornato", "")
        history_loaded = True


    data_list = []
    row_count = 0
    for row in rows:
        cols = row.find_all("td", recursive=False)
        if len(cols) < 5: continue # Salta righe vuote o spurie

        row_count += 1

        # --- Estrazione Titolo e Link ---
        link_tag = cols[2].find("a")
        titolo_testo = link_tag.text.strip()
        link_url = link_tag.get("href")
        link_url = base_url + "/" + link_url

        if enable_details_page_scraping:
            print(f"Scraping dettagli per evento [{row_count} / {len(rows) - 1}]: {titolo_testo}")
            # Estrazione dettagli dalla pagina dell'evento
            detail_response = requests.get(link_url, headers=headers)
            detail_soup = BeautifulSoup(detail_response.text, 'html.parser')
            
            apertura_iscr_tag = detail_soup.find("span", {"id": "MainContent_EventFormView_lbSubsFrom"})
            apertura_iscr = clean_data(apertura_iscr_tag.text.strip())

            chiusura_iscr_tag = detail_soup.find("span", {"id": "MainContent_EventFormView_lbSubsTo"})
            chiusura_iscr = clean_data(chiusura_iscr_tag.text.strip())
        
        if tipo_eventi == "Competenza":
            # gli eventi competenza riportano come "regione" sempre "nazionale", la regione è riportata nel titolo
            regione = re.search(r"Competenza - ([^|]+)", titolo_testo).group(1).strip()
        else:
            regione = cols[3].find("span").text.strip()


        partenza = clean_data(cols[4].text)
        rientro = clean_data(cols[5].text)

        quota_raw = cols[6].text.strip()
        try:
            quota_clean = quota_raw.replace('€', '').replace(',', '.').strip()
            quota = float(quota_clean)
        except ValueError:
            quota = 0.0  # Gestione errore se il campo è vuoto o contiene testo

        # --- Estrazione Località (Comune e Provincia) ---
        # Assume pattern "Nome Comune (PR)"
        localita_raw = cols[7].text.strip()
        match_loc = re.search(r"(.+)\s\((.+)\)", localita_raw)
        comune = match_loc.group(1).strip() if match_loc else localita_raw
        provincia = match_loc.group(2).strip() if match_loc else ""

        # --- Estrazione Iscritti (Pattern "10 / 20") ---
        iscritti_raw = cols[8].text.strip()
        match_iscr = re.search(r"(\d+)\s*/\s*(\d+)", iscritti_raw)
        iscritti_val = int(match_iscr.group(1)) if match_iscr else 0
        iscritti_max = int(match_iscr.group(2)) if match_iscr else 0

        # Creazione del record
        record = {
            "Titolo": titolo_testo,
            "Link": link_url,
            "Regione": regione,
            "Partenza": partenza,
            "Rientro": rientro,
            "Quota": quota,
            "Comune": comune,
            "Provincia": provincia,
            "Iscritti": iscritti_val,
            "Iscritti_MAX": iscritti_max,
            "Stato": "LIBERO" if iscritti_val < iscritti_max else ("CODA" if iscritti_val <= iscritti_max + 5 else "PIENO"),
            "Nuovo": False if history_loaded and titolo_testo in titoli_old else True
        }

        if enable_details_page_scraping:
            record["Apertura_Iscrizioni"] = apertura_iscr
            record["Chiusura_Iscrizioni"] = chiusura_iscr

        data_list.append(record)

    # --- CREAZIONE DATAFRAME ---
    df = pd.DataFrame(data_list)

    # --- EXPORT JSON ---
    output_json = {
        "tipo_eventi": tipo_eventi,
        "aggiornato": datetime.now().isoformat(),
        "aggiornamento_precedente": aggiornamento_precedente if history_loaded else "",
        "eventi": df.to_dict(orient="records")
    }

    def json_serial(obj):
        if isinstance(obj, (datetime)): return obj.strftime('%Y-%m-%d')
        raise TypeError ("Type %s not serializable" % type(obj))

    with open(json_filename, "w", encoding="utf-8") as f:
        json.dump(output_json, f, indent=4, default=json_serial, ensure_ascii=False)

    # --- PREPARAZIONE EXCEL ---

    # 1. Inizializza il Workbook
    wb = Workbook()
    ws = wb.active
    ws.title = worsheet_title

    # 2. Definisci l'intestazione (senza la colonna "Link" che integreremo nel Titolo)
    headers = ["Titolo", "Regione", "Provincia", "Comune", "Partenza", "Rientro", "Quota", "Iscritti", "Iscritti_MAX", "Stato", "Nuovo"]
    if enable_details_page_scraping:
        headers.extend(["Apertura Iscr", "Chiusura Iscr"])
    ws.append(headers)

    oggi = datetime.now()
        
    # file_excel_old_exists = False
    # path_excel = Path(file_excel)
    # path_excel_old = path_excel.with_name(path_excel.stem + "_old" + path_excel.suffix)
    # if path_excel.is_file():
    #     file_excel_old_exists = True
    #     shutil.copy(path_excel, path_excel_old) # Backup del file esistente
    #     wb_old = load_workbook(path_excel_old, read_only=True)
    #     ws_old = wb_old[worsheet_title]


    # def exists_in_worksheet(ws, value):
    #     found = False
    #     for row in ws.iter_rows(min_row=2, min_col=1, max_col=1, values_only=True):
    #         if row[0] == value:
    #             found = True
    #             break
                
    #     return found

    # 3. Itera sulla lista di record e scrivi le righe
    for i, record in enumerate(data_list, start=2): # Start=2 perché la riga 1 è l'header
        
        cell_titolo = ws.cell(row=i, column=1, value=record["Titolo"])
        cell_titolo.font = Font(color="0000FF", underline="single")
        cell_titolo.hyperlink = record["Link"]
            

        cell_regione = ws.cell(row=i, column=2, value=record["Regione"])
        if 'regioni_filtro' in locals() and record["Regione"] not in regioni_filtro:
            ws.row_dimensions[i].hidden = True

        cell_provincia = ws.cell(row=i, column=3, value=record["Provincia"])
        cell_provincia.alignment = Alignment(horizontal='center', vertical='center')

        cell_comune = ws.cell(row=i, column=4, value=record["Comune"])
        cell_comune.alignment = Alignment(horizontal='left', vertical='center')

        cell_partenza = ws.cell(row=i, column=5, value=record["Partenza"])
        cell_partenza.number_format = 'DD/MM/YYYY'
        cell_partenza.alignment = Alignment(horizontal='center', vertical='center')

        cell_rientro = ws.cell(row=i, column=6, value=record["Rientro"])
        cell_rientro.number_format = 'DD/MM/YYYY'
        cell_rientro.alignment = Alignment(horizontal='center', vertical='center')


        cell_quota = ws.cell(row=i, column=7, value=record["Quota"])
        cell_quota.number_format = '#,##0.00 €'
        cell_quota.alignment = Alignment(horizontal='center', vertical='center')

        cell_iscritti = ws.cell(row=i, column=8, value=record["Iscritti"])
        cell_iscritti.alignment = Alignment(horizontal='center', vertical='center')
        cell_iscritti_max = ws.cell(row=i, column=9, value=record["Iscritti_MAX"])
        cell_iscritti_max.alignment = Alignment(horizontal='center', vertical='center')

        cell_stato = ws.cell(row=i, column=10, value=f'=IF(H{i}<I{i},"LIBERO",IF(H{i}<=(I{i}+5),"CODA","PIENO"))')
        cell_stato.alignment = Alignment(horizontal='center', vertical='center')

        if 'stato_filtro' in locals():
            if record["Iscritti"] < record["Iscritti_MAX"] and "LIBERO" not in stato_filtro:
                ws.row_dimensions[i].hidden = True
            elif record["Iscritti"] <= record["Iscritti_MAX"]+5 and "CODA" not in stato_filtro:
                ws.row_dimensions[i].hidden = True
            elif record["Iscritti"] > record["Iscritti_MAX"]+5 and "PIENO" not in stato_filtro:
                ws.row_dimensions[i].hidden = True

        if record["Nuovo"]:
            cell_nuovo = ws.cell(row=i, column=11, value="NUOVO")
            cell_nuovo.alignment = Alignment(horizontal='center', vertical='center')
        else:
            cell_nuovo = ws.cell(row=i, column=11, value="")
            cell_nuovo.alignment = Alignment(horizontal='center', vertical='center')

        if enable_details_page_scraping:
            cell_apertura = ws.cell(row=i, column=12, value=record["Apertura_Iscrizioni"])
            cell_apertura.number_format = 'DD/MM/YYYY'
            cell_apertura.alignment = Alignment(horizontal='center', vertical='center')

            cell_chiusura = ws.cell(row=i, column=13, value=record["Chiusura_Iscrizioni"])
            cell_chiusura.number_format = 'DD/MM/YYYY'
            cell_chiusura.alignment = Alignment(horizontal='center', vertical='center')

            if record["Chiusura_Iscrizioni"] < oggi:
                ws.row_dimensions[i].hidden = True




    ws.column_dimensions['A'].width = 60 # Titolo con link, serve più spazio
    ws.column_dimensions['B'].width = 12 # Regione
    ws.column_dimensions['C'].width = 10 # Provincia
    ws.column_dimensions['D'].width = 12 # Comune
    ws.column_dimensions['E'].width = 10 # Partenza
    ws.column_dimensions['F'].width = 10 # Rientro
    ws.column_dimensions['G'].width = 10 # Quota
    ws.column_dimensions['H'].width = 12 # Iscritti
    ws.column_dimensions['I'].width = 12 # Iscritti_MAX
    ws.column_dimensions['J'].width = 12 # Stato
    ws.column_dimensions['K'].width = 12 # Nuovo

    if enable_details_page_scraping:
        ws.column_dimensions['L'].width = 12 # Apertura Iscrizioni
        ws.column_dimensions['M'].width = 12 # Chiusura Iscrizioni

    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.fill = PatternFill(start_color='D9E1F2', end_color='D9E1F2', fill_type='solid') # Azzurro chiaro

    ws.cell(row=1, column=1).alignment = Alignment(horizontal='left', vertical='center')
    ws.cell(row=1, column=6).alignment = Alignment(horizontal='left', vertical='center')

    rosso = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid') # Rosso chiaro
    giallo = PatternFill(start_color='FFEB9C', end_color='FFEB9C', fill_type='solid') # Giallo chiaro
    verde = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid') # Verde chiaro
    fucsia = PatternFill(start_color='9975BB', end_color='9975BB', fill_type='solid')

    # 2. Definiamo il range (es: dalla riga 2 alla fine della colonna J)
    range_stato = f'J2:J{ws.max_row}'

    # 3. Applichiamo le regole
    ws.conditional_formatting.add(range_stato, CellIsRule(operator='equal', formula=['"PIENO"'], fill=rosso))
    ws.conditional_formatting.add(range_stato, CellIsRule(operator='equal', formula=['"CODA"'], fill=giallo))
    ws.conditional_formatting.add(range_stato, CellIsRule(operator='equal', formula=['"LIBERO"'], fill=verde))
    ws.conditional_formatting.add(f'K2:K{ws.max_row}', CellIsRule(operator='equal', formula=['"NUOVO"'], fill=fucsia))

    # 2. Attiviamo il filtro
    ws.auto_filter.ref = ws.dimensions
    ws.freeze_panes = "A2"

    if 'regioni_filtro' in locals():
        ws.auto_filter.add_filter_column(1, regioni_filtro)
    if 'stato_filtro' in locals():
        ws.auto_filter.add_filter_column(9, stato_filtro)


    # aggiornamento del filtro
    flt_col = FilterColumn(colId=12) # zero-based index
    c_filters = CustomFilters()
    c_filters.customFilter.append(CustomFilter(operator='greaterThanOrEqual', val=str(to_excel(oggi.replace(hour=0, minute=0, second=0, microsecond=0)))))
    flt_col.customFilters = c_filters
    ws.auto_filter.filterColumn.append(flt_col)

    # 5. Salvataggio
    wb.save(file_excel)
    print(f"File {file_excel} creato correttamente.")

    # if path_excel_old.is_file():
    #     wb_old.close()
    #     path_excel_old.unlink()



if __name__ == "__main__":
    main()